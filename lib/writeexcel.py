#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'YinJia'

import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import RED,GREEN,DARKYELLOW
import configparser as cparser

# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG,encoding='UTF-8')
name = cf.get("tester","name")

class WriteExcel():
    """文件写入数据"""
    def __init__(self,fileName):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE,setting.TARGET_FILE)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self,row_n,value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
        self.ws.cell(row_n, 13, name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)

    def write_data(self, row, column, value, way=None):
        """
        向excel中写入值

        :author: 王波
        :param row: 单元格行数，从1开始，int类型
        :param column: 单元格列数，从1开始，int类型
        :param value: 写入单元格中的值
        :param way: 写入单元格中的方式，way=None，表示默认操作，覆盖原来单元格中的值；way=append，表示追加
        :return:
        """
        str_value = str(value)
        if way:
            self.ws.cell(row, column, str_value)
        else:
            old_cell_value = self.ws.cell(row, column).value + "\n" + str_value
            self.ws.cell(row, column, old_cell_value)
        self.wb.save(self.filename)


# if __name__ == "__main__":
    # fileName = "../report/excelReport/DemoPharmacistAPITestCase.xlsx"
    # print(openpyxl.__version__)
    # we = WriteExcel(fileName)
    # we.write_data(2, 12, "hi zhouyu")